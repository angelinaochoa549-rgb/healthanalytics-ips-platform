import io
import csv
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from apps.etl.models import Paciente, ETLLog


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_csv(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="pacientes_clinicos.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'ID Paciente', 'Nombres', 'Apellidos', 'Edad', 'Sexo',
        'Peso', 'Altura', 'IMC', 'Clasificación IMC',
        'Presión Sistólica', 'Presión Diastólica', 'Frecuencia Cardíaca',
        'Glucosa', 'Colesterol', 'Saturación O2', 'Temperatura',
        'Fumador', 'Alcohol', 'Actividad Física',
        'Diagnóstico', 'Riesgo', 'Es Crítico', 'Fecha Consulta',
    ])

    for p in Paciente.objects.all():
        writer.writerow([
            p.id_paciente, p.nombres, p.apellidos, p.edad, p.sexo,
            p.peso, p.altura, p.imc, p.clasificacion_imc,
            p.presion_sistolica, p.presion_diastolica, p.frecuencia_cardiaca,
            p.glucosa, p.colesterol, p.saturacion_oxigeno, p.temperatura,
            p.fumador, p.consumo_alcohol, p.actividad_fisica,
            p.diagnostico_preliminar, p.riesgo_enfermedad,
            p.es_critico, p.fecha_consulta,
        ])
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pacientes Clínicos'

    headers = [
        'ID Paciente', 'Nombres', 'Apellidos', 'Edad', 'Sexo',
        'IMC', 'Clasif. IMC', 'Presión Sist.', 'Glucosa',
        'Colesterol', 'Diagnóstico', 'Riesgo', 'Es Crítico', 'Fecha Consulta',
    ]

    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    riesgo_colors = {
        'bajo': 'C8E6C9', 'medio': 'FFF9C4',
        'alto': 'FFCCBC', 'critico': 'FFCDD2',
    }

    for row, p in enumerate(Paciente.objects.all(), 2):
        data = [
            p.id_paciente, p.nombres, p.apellidos, p.edad, p.sexo,
            p.imc, p.clasificacion_imc, p.presion_sistolica, p.glucosa,
            p.colesterol, p.diagnostico_preliminar, p.riesgo_enfermedad,
            p.es_critico, str(p.fecha_consulta),
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

        color = riesgo_colors.get(p.riesgo_enfermedad, 'FFFFFF')
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=color, end_color=color, fill_type='solid'
            )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pacientes_clinicos.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph('Reporte Clínico — HealthAnalytics IPS', styles['Title']))
    elements.append(Spacer(1, 12))

    data = [['ID', 'Nombres', 'Apellidos', 'Edad', 'IMC', 'Glucosa', 'Riesgo', 'Crítico', 'Fecha']]
    for p in Paciente.objects.all()[:500]:
        data.append([
            p.id_paciente, p.nombres[:15], p.apellidos[:15],
            p.edad, p.imc, round(p.glucosa, 1),
            p.riesgo_enfermedad, 'Sí' if p.es_critico else 'No',
            str(p.fecha_consulta),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3c5e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_clinico.pdf"'
    return response